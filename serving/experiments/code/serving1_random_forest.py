import joblib
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import csv
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_log_error
from sklearn.model_selection import train_test_split, RandomizedSearchCV
from sklearn.preprocessing import MinMaxScaler

# Set model
model_name = "random_forest"

# Load data
dataset = pd.read_csv('./data/cpu_quota.csv', names=['thread_quota', 'packet_size', 'bandwidth_tx', 'pps_tx', 'cpu_usage'])
y = np.array(dataset['thread_quota'])
X = np.array(dataset.drop('thread_quota', axis=1))

# Split data
train_X, test_X, train_y, test_y = train_test_split(X, y, test_size=0.20, random_state=40)

# Generate xlsx sheet and input data
# wb = load_workbook('./data/data.xlsx')
# ws = wb["result"]
# ws.cell(1, 1).value = 'packet_size'
# ws.cell(1, 2).value = 'bandwidth_tx'
# ws.cell(1, 3).value = 'pps_tx'
# ws.cell(1, 4).value = 'cpu_usage'
# for j in range(100):
#     ws.cell(j+2, 1).value = test_X[j][0]
#     ws.cell(j+2, 2).value = test_X[j][1]
#     ws.cell(j+2, 3).value = test_X[j][2]
#     ws.cell(j+2, 4).value = test_X[j][3]
# wbb = load_workbook('./data/result.xlsx')
# wss = wbb["result"]
# wss.cell(1, 1).value = 'network_throughput'
# wss.cell(1, 2).value = 'SLO'

# Preprocess data
min_max_scalar = MinMaxScaler()
train_X_ppr = min_max_scalar.fit_transform(train_X)
test_X_ppr = min_max_scalar.transform(test_X)
train_y_ppr = min_max_scalar.fit_transform(train_y.reshape(-1, 1))
test_y_ppr = min_max_scalar.transform(test_y.reshape(-1, 1))

# Predict
clf = joblib.load("./model/" + model_name)
pred_y_ppr = clf.predict(test_X_ppr)
pred_y = min_max_scalar.inverse_transform(pred_y_ppr.reshape(-1, 1))

# Save
# ws.cell(1, 5).value = 'cpu_quota'
# for j in range(100):
#     ws.cell(j+2, 5).value = pred_y[j][0]
# wb.save('./data/data.xlsx')

with open('./data/input_raw.csv','w') as f:
    makewrite = csv.writer(f)
    makewrite.writerow(['packet_size', 'bandwidth_tx', 'pps_tx', 'cpu_usage', 'cpu_quota'])
    for i in range(100):
        makewrite.writerow([test_X[i][0], test_X[i][1], test_X[i][2], test_X[i][3], pred_y[i][0]])