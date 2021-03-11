import joblib
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestRegressor
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split, RandomizedSearchCV
from sklearn.preprocessing import MinMaxScaler

# Set model
model_name = "random_forest"

# Load data
dataset = pd.read_csv('../../data/cpu_quota.csv', names=['thread_quota', 'packet_size', 'bandwidth_tx', 'pps_tx', 'cpu_usage'])
y = np.array(dataset['thread_quota'])
X = np.array(dataset.drop('thread_quota', axis=1))

# Split data
train_X, test_X, train_y, test_y = train_test_split(X, y, test_size=0.20, random_state=40)

# Preprocess data
min_max_scalar = MinMaxScaler()
train_X_ppr = min_max_scalar.fit_transform(train_X)
test_X_ppr = min_max_scalar.transform(test_X)
train_y_ppr = min_max_scalar.fit_transform(train_y.reshape(-1, 1))
test_y_ppr = min_max_scalar.transform(test_y.reshape(-1, 1))

# Cross validate
n_estimators = list(range(100, 1000, 10))
max_features = [1, 2, 3, 4, 5]
min_samples_split = [2, 5, 10]
min_samples_leaf = [1, 2, 4]
random_grid = {'n_estimators': n_estimators,
               'max_features': max_features,
               'min_samples_split': min_samples_split,
               'min_samples_leaf': min_samples_leaf}

clf = RandomForestRegressor(random_state=40)
clf_random_cv = RandomizedSearchCV(estimator=clf, param_distributions=random_grid, n_iter=100, cv=5, scoring='neg_root_mean_squared_error', verbose=2, n_jobs=-1, random_state=40)
clf_random_cv.fit(train_X_ppr, train_y_ppr.ravel())

# Save model
joblib.dump(clf_random_cv, "../../model/cpu_quota_rmse/" + model_name)

# Test
clf = joblib.load("../../model/cpu_quota_rmse/" + model_name)
pred_y_ppr = clf.predict(test_X_ppr)
pred_y = min_max_scalar.inverse_transform(pred_y_ppr.reshape(-1, 1))

# Evaluate
score = mean_squared_error(test_y, pred_y, squared=False)

# Save
wb = load_workbook('../../data/cpu_quota_rmse.xlsx')

ws = wb["results"]
ws.cell(3, 8).value = model_name
ws.cell(3, 9).value = score

ws = wb[model_name]
ws.cell(1, 1).value = 'pred_y'
ws.cell(1, 2).value = 'test_y'
for j in range(100):
    ws.cell(j+2, 1).value = pred_y[j][0]
for j in range(100):
    ws.cell(j+2, 2).value = test_y[j]

wb.save('C:/Users/Jiyou/Desktop/github.com/ksc-2020/data/cpu_quota_rmse.xlsx')
