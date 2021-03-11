import joblib
from openpyxl import load_workbook
import numpy as np
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.model_selection import train_test_split, RandomizedSearchCV
from sklearn.preprocessing import MinMaxScaler

# Set model
from sklearn.svm import SVR

model_name = "support_vector"

# Load data
dataset = pd.read_csv('../../data/cpu_usage.csv', names=['thread_quota', 'packet_size', 'bandwidth_tx', 'pps_tx', 'cpu_usage'])
y = np.array(dataset['cpu_usage'])
X = np.array(dataset.drop('cpu_usage', axis=1))

# Split data
train_X, test_X, train_y, test_y = train_test_split(X, y, test_size=0.20, random_state=40)

# Preprocess data
min_max_scalar = MinMaxScaler()
train_X_ppr = min_max_scalar.fit_transform(train_X)
test_X_ppr = min_max_scalar.transform(test_X)
train_y_ppr = min_max_scalar.fit_transform(train_y.reshape(-1, 1))
test_y_ppr = min_max_scalar.transform(test_y.reshape(-1, 1))

# # Cross validate
# C = list(range(10, 2000, 10))
# degree = list(range(1, 5, 1))
# random_grid = {'C': C,
#                'degree': degree}
#
# clf = SVR(kernel='poly')
# clf_random_cv = RandomizedSearchCV(estimator=clf, param_distributions=random_grid, n_iter=100, cv=5, scoring='neg_root_mean_squared_error', verbose=2, n_jobs=-1, random_state=40)
# clf_random_cv.fit(train_X_ppr, train_y_ppr.ravel())
#
# # Save model
# joblib.dump(clf_random_cv, "../../model/cpu_usage/" + model_name)

# Test
clf = joblib.load("../../model/cpu_usage/" + model_name)
pred_y_ppr = clf.predict(test_X_ppr)
pred_y = min_max_scalar.inverse_transform(pred_y_ppr.reshape(-1, 1))

# Evaluate
score = mean_squared_error(test_y, pred_y, squared=False)

# Save
wb = load_workbook('../../data/cpu_usage.xlsx')

ws = wb["results"]
ws.cell(4, 8).value = model_name
ws.cell(4, 9).value = score

ws = wb[model_name]
ws.cell(1, 1).value = 'pred_y'
ws.cell(1, 2).value = 'test_y'
for j in range(100):
    ws.cell(j+2, 1).value = pred_y[j][0]
for j in range(100):
    ws.cell(j+2, 2).value = test_y[j]

wb.save('C:/Users/Jiyou/Desktop/github.com/ksc-2020/data/cpu_usage.xlsx')
